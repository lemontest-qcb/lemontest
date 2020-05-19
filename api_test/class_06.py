#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/2/25 21:53
# @Author : muji
# @QQ: 2378807189
# Copyright：湖南省零檬信息技术有限公司

# 操作excel表格：读取数据+ 回写数据（结果） --- 第三方库： openpyxl
# 1、 安装    pip install openpyxl
# 2、 导入
'''
excel表格的三大对象：
1、工作簿对象：文件名称
2、 表单对象： sheet
3、表格对象： cell
cell = sheet.cell(row=1,column=1)  #取到表格对象
print(cell.value)  #获取到表格里的内容
cell.value = "测试编号" # 给这个表格内容重新赋值  --写入--切记：一定要保存
print(cell.value)
wb.save("test_case.xlsx") # 保存  -- 先关闭文件
'''
'''
1、 写测试用例（手工）----> 自动化接口测试用例  ==done
2、 代码自动读取测试数据 --- done
3、 数据发送接口请求  --执行结果   == done
4、 执行结果 预期结果 做断言---执行是否通过？---- 结论
5、 通过---标记passed；不通过--- bug+failed。--回写结果excel表格
'''
import openpyxl # 导入库
import requests
session = requests.session() #模块付给一个变量
#发送接口请求函数
def post_fun(url,data):
    # response = requests.post(url,data)
    response = session.post(url,data)  #直接带上cookies--先登录
    result = response.json() # 获取相应结果--字典
    return result
#读取数据的函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载为工作簿对象
    sheet = wb[sheetname]  # 表单
    max_row = sheet.max_row #获取当前表单的最大行号
    case_list = [] #定义一个空列表，用来存放所有的测试用例的（字典）
    for i in range(2,max_row+1):
        case = dict(
        case_id =  sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value, # 读取到了url
        data = sheet.cell(row=i,column=6).value, # 参数
        expected_result = sheet.cell(row=i,column=7).value) #预期结果
        case_list.append(case) #往空列表里追加元素--字典
    return case_list #返回值---保存了所有测试用例的列表
#回写结果的函数
def write_result(filename,sheetname,row,column,result):
    wb = openpyxl.load_workbook(filename)  # 加载为工作簿对象
    sheet = wb[sheetname]
    sheet.cell(row,column).value = result #回写数据
    wb.save(filename) #保存文件
#{'case_id': 1,
# 'url': 'http://test.lemonban.com/futureloan/mvc/api/member/register',
# 'data': "{'mobilephone':'13732371298','pwd':'123456','regname':'meson'}",
# 'expected_result': '{"status":1,"code":"10001","data":null,"msg":"注册成功"}'}
# {'status': 0, 'code': '20110', 'data': None, 'msg': '手机号码已被注册'}
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname) #调用了读取数据的函数
    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")
        data = case["data"]  #是字符串--不能直接给requests.post传参--字典
        data = eval(data) # 转化为字典
        expected_result = case.get("expected_result") #预期结果
        expected_result = expected_result.replace("null","None") #替换
        expected_result = eval(expected_result) #转化为字典
        expected_msg = expected_result.get("msg") #预期结果取值
        real_result = post_fun(url,data) #调用发送接口请求的函数
        real_msg = real_result["msg"] # 得到断言的字段-字典取值
        print('预期结果是：{}'.format(expected_msg))
        print('真实结果是：{}'.format(real_msg))
        if expected_msg == real_msg:
            print("第{}条测试用例通过的！".format(case_id))
            final_result = "Passed"
        else:
            print("第{}条测试用例不通过！".format(case_id))
            final_result = "Failed"
        print("*" * 20)
        write_result(filename,sheetname,case_id+1,8,final_result) #调用回写数据的函数

execute_fun("test_case.xlsx","login")
# execute_fun("test_case.xlsx","recharge")
















